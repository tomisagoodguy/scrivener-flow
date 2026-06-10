import pathlib
import pandas as pd
import re
import logging
from openpyxl import load_workbook
from typing import Optional, Tuple, Dict, Any

from ETF.utils.date_parser import normalize_date

logger = logging.getLogger(__name__)

# Constants for Magic Strings
KEY_DATA_DATE: str = "資料日期"
KEY_STOCK_NAME: str = "股票名稱"
KEY_STOCK_CODE: str = "股票代號"
KEY_SHARES: str = "股數"
KEY_WEIGHT: str = "權重"
KEY_RATIO: str = "比例"

COL_CODE: str = "code"
COL_NAME: str = "name"
COL_SHARES: str = "shares"
COL_WEIGHT: str = "weight"

# Alias mapping: standard column name → all known header variants across issuers
COLUMN_ALIASES: dict[str, list[str]] = {
    COL_CODE: ["證券代號", "股票代號", "代號", "代碼", "stock_code", "StockCode"],
    COL_NAME: ["證券名稱", "股票名稱", "名稱", "stock_name", "StockName"],
    COL_WEIGHT: ["持股比重", "比重", "權重", "持股比例", "weight_pct", "Weight"],
    COL_SHARES: ["持有股數", "持有數量", "股數", "持股數", "shares_held", "Shares"],
}

def _clean_shares(x: Any) -> int:
    try:
        return int(float(str(x).replace(",", "")))
    except (ValueError, TypeError):
        return 0

def _clean_weight(x: Any) -> float:
    try:
        return float(str(x).replace("%", "").replace(",", ""))
    except (ValueError, TypeError):
        return 0.0

def _extract_date_from_sheet(ws: Any) -> Optional[str]:
    """
    Scan the top-left area of the worksheet to find the data date.
    """
    for r in range(1, 10):  # Scan top 10 rows
        for c in range(1, 6):  # Scan first 5 columns
            v = ws.cell(row=r, column=c).value
            s_val = str(v) if v else ""
            
            if KEY_DATA_DATE in s_val:
                # Case 1: "資料日期：2026/01/22" in one cell
                if ":" in s_val or "：" in s_val:
                    parts = re.split(r"[:：]", s_val)
                    if len(parts) > 1:
                        normalized = normalize_date(parts[1])
                        if normalized:
                            return normalized
                
                # Case 2: "資料日期" in cell A, Date in cell B (next column)
                next_v = ws.cell(row=r, column=c+1).value
                if next_v:
                    normalized = normalize_date(str(next_v))
                    if normalized:
                        return normalized
    return None

def _identify_header_row(df: pd.DataFrame) -> int:
    """Find the row index containing both Stock Name and Stock Code."""
    for i, row in df.iterrows():
        row_str = " ".join(row.fillna("").astype(str))
        if KEY_STOCK_NAME in row_str and KEY_STOCK_CODE in row_str:
            return int(i)  # type: ignore[arg-type]
    return -1

def _standardize_columns(cols: pd.Index) -> Dict[str, str]:
    """Map raw column names to standardized keys via COLUMN_ALIASES.

    Priority: exact match on standard name → exact match on alias → log warning and exclude.
    Unknown columns are excluded from the result (no exception raised).
    """
    # Build reverse lookup: alias (and standard name itself) → standard name
    alias_to_std: dict[str, str] = {}
    for std_name, aliases in COLUMN_ALIASES.items():
        alias_to_std[std_name] = std_name
        for alias in aliases:
            alias_to_std[alias] = std_name

    col_map: Dict[str, str] = {}
    for c in cols:
        c_str = str(c).strip()
        if c_str in alias_to_std:
            col_map[c] = alias_to_std[c_str]
        else:
            logger.warning(f"Unknown column '{c_str}', excluded from result")
    return col_map

def parse_holdings_xlsx(xlsx_path: pathlib.Path) -> Tuple[pd.DataFrame, Optional[str]]:
    """
    Parse 00981A specific holding XLSX.
    Returns (DataFrame, data_date_str)
    """
    logger.info(f"Parsing XLSX: {xlsx_path}")
    
    # 1. Extract Data Date
    data_date = None
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            if wb.worksheets:
                data_date = _extract_date_from_sheet(wb.worksheets[0])
            else:
                logger.warning("Workbook has no worksheets.")
        finally:
            wb.close()
    except Exception as e:
        logger.error(f"Error extracting date: {e}")
        
    date_formatted = ""
    if data_date:
        # Format YYYYMMDD to YYYY-MM-DD
        if len(data_date) == 8:
             date_formatted = f"{data_date[:4]}-{data_date[4:6]}-{data_date[6:]}"
        else:
             date_formatted = data_date  # Should align with downstream expectation
    else:
        logger.warning("Could not find Data Date in Excel content.")
    
    # 2. Extract Holdings
    try:
        # Read headerless first to locate real header
        df_raw = pd.read_excel(xlsx_path, header=None, dtype=str)
        
        header_idx = _identify_header_row(df_raw)
        if header_idx == -1:
            logger.error(f"Could not find header row with '{KEY_STOCK_NAME}' and '{KEY_STOCK_CODE}'")
            return pd.DataFrame(), date_formatted
            
        # Set columns and slice data
        df_raw.columns = df_raw.iloc[header_idx]
        df = df_raw.iloc[header_idx+1:].copy()
        
        # Standardize columns
        col_map = _standardize_columns(df.columns)
        df = df.rename(columns=col_map)
        
        required = [COL_CODE, COL_NAME, COL_SHARES, COL_WEIGHT]
        missing = [col for col in required if col not in df.columns]
        if missing:
            logger.error(f"Missing required columns: {missing}. Found: {df.columns.tolist()}")
            return pd.DataFrame(), date_formatted
            
        # Clean Data
        df = df.dropna(subset=[COL_CODE])
        df = df[df[COL_CODE].str.strip() != ""] # Remove empty codes
        
        # enforce types
        df[COL_CODE] = df[COL_CODE].astype(str).str.strip()
        df[COL_NAME] = df[COL_NAME].astype(str).str.strip()
        df[COL_SHARES] = df[COL_SHARES].apply(_clean_shares)
        df[COL_WEIGHT] = df[COL_WEIGHT].apply(_clean_weight)
        
        # Filter (remove summary rows or zero shares)
        df = df[df[COL_SHARES] > 0]
        
        return df, date_formatted

    except Exception as e:
        logger.error(f"Error parsing holdings: {e}")
        return pd.DataFrame(), date_formatted
