"""
官網 API Scraper — 投信官方 API 備援爬蟲

CATALOG 來源：reference/tw-active/tools/etfdaily.py（Round 45 破解）

支援 ETF：
  統一 ezmoney      : 00981A, 00988A, 00403A (GET XLSX + cookie jar)
  野村 nomura       : 00980A, 00985A  (POST JSON)
  復華 fhtrust      : 00991A          (GET XLSX)
  安聯 allianz      : 00984A, 00993A  (POST JSON + ASP.NET antiforgery)
  群益 capital      : 00982A, 00992A, 00997A (POST JSON)
  元大 yuanta       : 00990A          (Playwright + window.__NUXT__)
  台新 taishin      : 00987A          (GET HTML + BeautifulSoup)
  第一金 first_financial : 00994A    (POST JSON double-decode)
  中信 ctbc         : 00995A          (POST JSON two-step auth token)
  兆豐 mega         : 00986A, 00996A  (GET HTML + BeautifulSoup, fund_id required)
  摩根 jpm          : 00401A, 00989A  (GET XLSX + openpyxl two-section parse)
  國泰 cathay       : 00400A          (GET JSON REST API)
  中信 HTML ctbc_html: 00983A         (GET HTML + BeautifulSoup, legacy ASP.NET page)

Public interface:
  fetch_holdings(etf_code, date_str=None) -> pd.DataFrame
    columns: code(str), name(str), weight(float,%), shares(int)
    empty DataFrame on any error (never raises)
"""

from __future__ import annotations

import io
import json
import logging
import re
import ssl
import urllib.error
import urllib.parse
import urllib.request
from http.cookiejar import CookieJar
from typing import Any, cast

import pandas as pd

# SSL context for sites with certificate issues (taishin, first_financial, ctbc etc.)
_SSL_NO_VERIFY = ssl.create_default_context()
_SSL_NO_VERIFY.check_hostname = False
_SSL_NO_VERIFY.verify_mode = ssl.CERT_NONE

logger = logging.getLogger(__name__)

UA = "Mozilla/5.0 (etf-pipeline-official-api; +scrivener-flow)"

# ── CATALOG ──────────────────────────────────────────────────────────────────

CATALOG: dict[str, dict[str, Any]] = {
    "00981A": {"issuer": "uni", "name": "統一台股增長", "fund_code": "49YTW"},
    "00988A": {"issuer": "uni", "name": "統一全球創新", "fund_code": "61YTW"},
    "00403A": {"issuer": "uni", "name": "統一升級50", "fund_code": "63YTW"},
    "00991A": {"issuer": "fhtrust", "name": "復華台灣未來50", "fund_code": "ETF23"},
    "00980A": {"issuer": "nomura", "name": "野村臺灣智慧優選", "fund_code": "00980A"},
    "00985A": {"issuer": "nomura", "name": "野村臺灣增強50", "fund_code": "00985A"},
    "00993A": {"issuer": "allianz", "name": "安聯台灣主動式", "fund_code": "E0002"},
    "00984A": {"issuer": "allianz", "name": "安聯台灣高息成長", "fund_code": "E0001"},
    "00982A": {"issuer": "capital", "name": "群益台灣精選強棒", "fund_code": "399"},
    "00992A": {"issuer": "capital", "name": "群益台灣科技創新", "fund_code": "500"},
    "00997A": {"issuer": "capital", "name": "群益美國增長", "fund_code": "502"},
    "00990A": {"issuer": "yuanta", "name": "元大AI新經濟", "fund_code": "00990A"},
    # ── 新增 4 家投信直接 API（stable-etf-scrapers）──
    "00987A": {"issuer": "taishin", "name": "台新優勢成長", "fund_code": "00987A"},
    "00986A": {"issuer": "taishin", "name": "台新龍頭成長", "fund_code": "00986A"},
    "00994A": {"issuer": "first_financial", "name": "第一金台股優", "fund_code": "182"},
    "00995A": {"issuer": "ctbc", "name": "中信台灣卓越", "fund_code": "E0036"},
    "00999A": {"issuer": "nomura", "name": "野村臺灣高息", "fund_code": "00999A"},
    # 兆豐：fund_id=23 由 TW_Active_Tracker 確認
    "00996A": {
        "issuer": "mega",
        "name": "兆豐台灣豐收",
        "fund_code": "00996A",
        "fund_id": "23",
    },
    # ── replace-pocket-scrapers：新增 4 支 ETF 直接 API ──
    "00400A": {"issuer": "cathay", "name": "國泰動能高息", "fund_code": "EA"},
    "00401A": {
        "issuer": "jpm",
        "name": "摩根台灣鑫收",
        "fund_code": "00401A",
        "xlsx_url": "https://am.jpmorgan.com/content/dam/jpm-am-aem/asiapacific/tw/zh/regulatory/etf-supplement/jpm_apac_tw_etf_pcf_updates_00401A_TW00000401A1.xlsx",
    },
    "00989A": {
        "issuer": "jpm",
        "name": "摩根美國科技",
        "fund_code": "00989A",
        "xlsx_url": "https://am.jpmorgan.com/content/dam/jpm-am-aem/asiapacific/tw/zh/regulatory/etf-supplement/jpm_apac_tw_etf_pcf_updates_00989A_TW00000989A5.xlsx",
    },
    "00983A": {"issuer": "ctbc_html", "name": "中信ARK創新", "fund_code": "00983A"},
    # ── expand-etf-coverage-and-diff-schema：新增 D 類 ETF ──
    # 聯博（00984D）：REST API，接受 ISIN 等非 4 碼代號
    "00984D": {
        "issuer": "alliance_bernstein",
        "name": "聯博全球非投",
        "fund_code": "TW00000984D0",
    },
    # 富邦（00982D / 00983D）：HTML 解析；兩檔均為債券型 ETF，預期無股票持股明細
    "00982D": {"issuer": "fubon", "name": "富邦動態入息", "fund_code": "00982D"},
    "00983D": {"issuer": "fubon", "name": "富邦複合收益", "fund_code": "00983D"},
}


def is_supported(etf_code: str) -> bool:
    return etf_code in CATALOG


# ── HTTP helpers ──────────────────────────────────────────────────────────────


def _get(
    url: str, cookie_jar: CookieJar | None = None, verify_ssl: bool = True
) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    ctx = None if verify_ssl else _SSL_NO_VERIFY
    if cookie_jar is not None:
        opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(cookie_jar)
        )
        with opener.open(req, timeout=30) as resp:
            return resp.read()
    with urllib.request.urlopen(req, timeout=30, context=ctx) as resp:
        return resp.read()


def _post_json(
    url: str,
    payload: dict[str, Any],
    extra_headers: dict[str, str] | None = None,
    cookie_jar: CookieJar | None = None,
    verify_ssl: bool = True,
) -> bytes:
    data = json.dumps(payload).encode("utf-8")
    headers = {
        "User-Agent": UA,
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    if extra_headers:
        headers.update(extra_headers)
    req = urllib.request.Request(url, data=data, headers=headers, method="POST")
    ctx = None if verify_ssl else _SSL_NO_VERIFY
    if cookie_jar is not None:
        opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(cookie_jar)
        )
        with opener.open(req, timeout=30) as resp:
            return resp.read()
    with urllib.request.urlopen(req, timeout=30, context=ctx) as resp:
        return resp.read()


# ── XLSX parser (統一 / 復華) ─────────────────────────────────────────────────


def _parse_xlsx(raw: bytes) -> list[dict[str, Any]]:
    """解析持股 XLSX，回傳 [{code, name, shares, weight_pct}, ...]"""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = list(ws.iter_rows(values_only=True))

    header_idx = -1
    for i, row in enumerate(rows):
        if not row:
            continue
        cells = [str(c).strip() if c is not None else "" for c in row]
        line = " ".join(cells)
        if ("代號" in line or "代碼" in line) and ("權重" in line or "比例" in line):
            header_idx = i
            break

    if header_idx < 0:
        return []

    holdings: list[dict[str, Any]] = []
    for row in rows[header_idx + 1 :]:
        if not row or row[0] is None:
            if holdings:
                break
            continue
        code = str(row[0]).strip()
        if not code:
            continue
        name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        shares = _to_num(row[2]) if len(row) > 2 else None
        weight = _to_num(row[-1]) if row else None
        holdings.append(
            {"code": code, "name": name, "shares": shares, "weight_pct": weight}
        )
    return holdings


def _to_num(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("%", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _fund_assets(
    aum: Any = None, nav: Any = None, units: Any = None, nav_date: Any = None
) -> dict[str, Any]:
    """組基金資產摘要 dict，單一欄位解析失敗即帶 None，不影響持股解析。

    aum/nav/units 經 _to_num 容錯（逗號、百分比、非數字 → None）；
    nav_date 為原始字串（空字串視為 None）。
    """
    date_str = str(nav_date).strip() if nav_date else ""
    return {
        "aum": _to_num(aum),
        "nav": _to_num(nav),
        "units": _to_num(units),
        "nav_date": date_str or None,
    }


def _fund_assets_or_none(
    aum: Any = None, nav: Any = None, units: Any = None, nav_date: Any = None
) -> dict[str, Any] | None:
    """同 `_fund_assets`，但 aum/nav/units 全為 None 時回傳 None（無摘要可用）。"""
    assets = _fund_assets(aum=aum, nav=nav, units=units, nav_date=nav_date)
    if assets["aum"] is None and assets["nav"] is None and assets["units"] is None:
        return None
    return assets


def _check_aum_nav_units_consistency(
    assets: dict[str, Any] | None, tolerance: float = 0.05
) -> dict[str, Any] | None:
    """aum ≈ nav × units 合理性檢查（誤差 >5% 時只留 nav，其餘設 None）。

    aum/nav/units 任一缺就跳過檢查（無法驗證，原樣回傳）。
    """
    if not assets:
        return assets
    aum, nav, units = assets.get("aum"), assets.get("nav"), assets.get("units")
    if aum is None or nav is None or units is None:
        return assets
    expected = nav * units
    if expected == 0:
        return assets
    err = abs(aum - expected) / abs(expected)
    if err > tolerance:
        logger.warning(
            "[FUND_ASSETS] aum(%.2f) 與 nav*units(%.2f) 誤差 %.1f%% > %.0f%%，"
            "僅保留 nav，aum/units 設 None",
            aum,
            expected,
            err * 100,
            tolerance * 100,
        )
        return {**assets, "aum": None, "units": None}
    return assets


def _xlsx_rows_to_text(rows: list[tuple[Any, ...]]) -> str:
    """把 XLSX rows 攤平成純文字（每列 cell 以空白相接），供 regex 抽取標籤數值。"""
    lines: list[str] = []
    for row in rows:
        if not row:
            continue
        cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if cells:
            lines.append(" ".join(cells))
    return "\n".join(lines)


def _parse_uni_assets(rows: list[tuple[Any, ...]]) -> dict[str, Any] | None:
    """統一 ezmoney XLSX 資產摘要：淨資產 / 流通在外單位數 / 每單位淨值 / 資料日期（民國年）。"""
    text = _xlsx_rows_to_text(rows)
    aum_m = re.search(r"淨資產\s*[:：]?\s*(?:NTD)?\s*([\d,]+)", text)
    units_m = re.search(r"流通在外單位數\s*[:：]?\s*([\d,]+)", text)
    nav_m = re.search(r"每單位淨值\s*[:：]?\s*(?:NTD)?\s*\$?\s*([\d,.]+)", text)
    date_m = re.search(r"資料日期\s*[:：]\s*(\d{2,3})/(\d{2})/(\d{2})", text)
    nav_date = None
    if date_m:
        year = int(date_m.group(1)) + 1911  # ROC → 西元
        nav_date = f"{year}-{date_m.group(2)}-{date_m.group(3)}"
    return _fund_assets_or_none(
        aum=aum_m.group(1) if aum_m else None,
        nav=nav_m.group(1) if nav_m else None,
        units=units_m.group(1) if units_m else None,
        nav_date=nav_date,
    )


def _parse_fhtrust_assets(rows: list[tuple[Any, ...]]) -> dict[str, Any] | None:
    """復華 XLSX 資產摘要：基金資產淨值 / 基金在外流通單位數 / 基金每單位淨值 / 日期（西元年）。"""
    text = _xlsx_rows_to_text(rows)
    aum_m = re.search(r"基金資產淨值\s*[:：]?\s*([\d,]+)", text)
    units_m = re.search(r"基金在外流通單位數\s*[:：]?\s*([\d,]+)", text)
    nav_m = re.search(r"基金每單位淨值\s*[:：]?\s*([\d,.]+)", text)
    date_m = re.search(r"日期\s*[:：]\s*(\d{4})/(\d{2})/(\d{2})", text)
    nav_date = (
        f"{date_m.group(1)}-{date_m.group(2)}-{date_m.group(3)}" if date_m else None
    )
    return _fund_assets_or_none(
        aum=aum_m.group(1) if aum_m else None,
        nav=nav_m.group(1) if nav_m else None,
        units=units_m.group(1) if units_m else None,
        nav_date=nav_date,
    )


# ── Per-issuer fetchers ───────────────────────────────────────────────────────


def _xlsx_raw_rows(raw: bytes) -> list[tuple[Any, ...]]:
    """讀 XLSX bytes，回傳 active sheet 的所有 rows（供持股+資產摘要共用一次解析）。"""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    if ws is None:
        return []
    return list(ws.iter_rows(values_only=True))


def _fetch_uni(
    fund_code: str,
) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
    jar = CookieJar()
    url = f"https://www.ezmoney.com.tw/ETF/Fund/AssetExcelNPOI?fundCode={fund_code}"
    raw = _get(url, cookie_jar=jar)
    if not raw.startswith(b"PK"):
        raw = _get(url, cookie_jar=jar)  # retry once after cookie warm-up
    if not raw.startswith(b"PK"):
        raise RuntimeError(f"統一 ezmoney 未回傳 XLSX，前 60 bytes: {raw[:60]!r}")
    holdings = _parse_xlsx(raw)
    # ezmoney XLSX 回傳千股（千株），轉換為股
    for h in holdings:
        if h.get("shares") is not None:
            h["shares"] = int(h["shares"] * 1000)
    assets = None
    try:
        rows = _xlsx_raw_rows(raw)
        assets = _check_aum_nav_units_consistency(_parse_uni_assets(rows))
    except Exception as exc:
        logger.error("[UNI] %s 資產摘要解析失敗（不影響持股）：%s", fund_code, exc)
    return holdings, assets


def _fetch_fhtrust(
    fund_code: str, date_ymd: str
) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
    url = f"https://www.fhtrust.com.tw/api/assetsExcel/{fund_code}/{date_ymd}"
    raw = _get(url)
    if raw.startswith(b"\xef\xbb\xbf") or b"\xe6\x9f\xa5" in raw[:20]:
        raise RuntimeError(f"復華 {date_ymd}：查無資料（非交易日或未發布）")
    if not raw.startswith(b"PK"):
        raise RuntimeError(f"復華未回傳 XLSX：{raw[:60]!r}")
    holdings = _parse_xlsx(raw)
    assets = None
    try:
        rows = _xlsx_raw_rows(raw)
        assets = _check_aum_nav_units_consistency(_parse_fhtrust_assets(rows))
    except Exception as exc:
        logger.error("[FHTRUST] %s 資產摘要解析失敗（不影響持股）：%s", fund_code, exc)
    return holdings, assets


def _fetch_nomura(
    fund_id: str, date_ymd: str | None
) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
    search_date = _ymd_to_dash(date_ymd) if date_ymd else _last_weekday_dash()
    payload = {"FundID": fund_id, "SearchDate": search_date}
    # 野村憑證鏈缺 Subject Key Identifier（2026-04-22 起），嚴格驗證必失敗
    # （certifi 也無效），比照 taishin/ctbc/fubon 停用驗證
    raw = _post_json(
        "https://www.nomurafunds.com.tw/API/ETFAPI/api/Fund/GetFundAssets",
        payload,
        verify_ssl=False,
    )
    js: dict[str, Any] = json.loads(raw.decode("utf-8"))
    data = (js.get("Entries") or {}).get("Data") or {}
    tables = data.get("Table") or []
    holdings: list[dict[str, Any]] = []
    for t in tables:
        title = (t.get("TableTitle") or "").strip()
        if not title:
            continue
        for row in t.get("Rows") or []:
            if len(row) < 4:
                continue
            raw_shares = _to_num(row[2])
            holdings.append(
                {
                    "code": str(row[0]).strip(),
                    "name": str(row[1]).strip(),
                    # Nomura API 回傳千股（千株），轉換為股
                    "shares": int(raw_shares * 1000)
                    if raw_shares is not None
                    else None,
                    "weight_pct": _to_num(row[3]),
                }
            )
    # 基金資產摘要：Entries.Data.FundAsset.{Aum,Nav,Units,NavDate}
    asset = data.get("FundAsset") or {}
    assets = (
        _fund_assets(
            aum=asset.get("Aum"),
            nav=asset.get("Nav"),
            units=asset.get("Units"),
            nav_date=asset.get("NavDate"),
        )
        if asset
        else None
    )
    return holdings, assets


def _fetch_allianz(
    fund_no: str, date_ymd: str | None
) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
    jar = CookieJar()
    _get(
        "https://etf.allianzgi.com.tw/webapi/api/AntiForgery/GetAntiForgeryToken",
        cookie_jar=jar,
    )
    token = next((c.value for c in jar if c.name == "X-XSRF-TOKEN"), "")
    if not token:
        raise RuntimeError("安聯：未取得 X-XSRF-TOKEN cookie")
    date_iso = (_ymd_to_dash(date_ymd) + "T00:00:00") if date_ymd else None
    payload = {"FundNo": fund_no, "Date": date_iso}
    raw = _post_json(
        "https://etf.allianzgi.com.tw/webapi/api/Fund/GetFundTradeInfo",
        payload,
        extra_headers={"X-XSRF-TOKEN": token},
        cookie_jar=jar,
    )
    js: dict[str, Any] = json.loads(raw.decode("utf-8"))
    entries = js.get("Entries") or {}
    tables = entries.get("DynamicTableData") or []
    holdings: list[dict[str, Any]] = []
    for t in tables:
        for row in t.get("Rows") or []:
            if not row or len(row) < 5:
                continue
            holdings.append(
                {
                    "code": str(row[1]).strip(),
                    "name": str(row[2]).strip(),
                    "shares": _to_num(row[3]),
                    "weight_pct": _to_num(row[4]),
                }
            )
    # 基金資產摘要：Entries.{CAnceTotalAv,CAnceNav,CAnceTotalIssues,CNavDt}
    assets = _fund_assets(
        aum=entries.get("CAnceTotalAv"),
        nav=entries.get("CAnceNav"),
        units=entries.get("CAnceTotalIssues"),
        nav_date=entries.get("CNavDt"),
    )
    if assets["aum"] is None and assets["nav"] is None and assets["units"] is None:
        assets = None
    return holdings, assets


def _fetch_capital(
    fund_id: str, date_ymd: str | None
) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
    payload: dict[str, Any] = {"fundId": fund_id}
    if date_ymd:
        payload["date"] = _ymd_to_dash(date_ymd)
    raw = _post_json("https://www.capitalfund.com.tw/CFWeb/api/etf/buyback", payload)
    js: dict[str, Any] = json.loads(raw.decode("utf-8"))
    data = js.get("data") or {}
    holdings: list[dict[str, Any]] = []
    for s in data.get("stocks") or []:
        holdings.append(
            {
                "code": str(s.get("stocNo", "")).strip(),
                "name": str(s.get("stocName", "")).strip(),
                "shares": _to_num(s.get("share")),
                "weight_pct": _to_num(s.get("weight")),
            }
        )
    # 基金資產摘要：data.pcf.{nav=淨資產, pUnit=每單位淨值, totUnit=流通單位, date1}
    pcf = data.get("pcf") or {}
    assets = (
        _fund_assets(
            aum=pcf.get("nav"),
            nav=pcf.get("pUnit"),
            units=pcf.get("totUnit"),
            nav_date=pcf.get("date1"),
        )
        if pcf
        else None
    )
    return holdings, assets


# ── 元大 (00990A) 資產摘要 — 輕量 requests，不需 Playwright ──────────────────────


def _fetch_yuanta_assets(fund_code: str) -> dict[str, Any] | None:
    """元大 ETF 資產摘要：伺服器端渲染 HTML 內即含 NAV，不需 Playwright/JS 執行。

    僅補資產摘要，持股解析仍走既有 yuanta_scraper（Playwright）。
    失敗回 None，不影響持股。
    """
    from bs4 import BeautifulSoup

    try:
        url = f"https://www.yuantaetfs.com/tradeInfo/pcf/{fund_code}"
        raw = _get(url, verify_ssl=False)
        html = raw.decode("utf-8", errors="replace")
        soup = BeautifulSoup(html, "lxml")
        text = soup.get_text(" ", strip=True)

        aum_m = re.search(r"基金淨資產價值\s*NTD\s*\$?\s*([\d,]+(?:\.\d+)?)", text)
        nav_m = re.search(
            r"每受益權單位淨資產價值\(元\)\s*NTD\s*\$?\s*([\d,]+(?:\.\d+)?)", text
        )
        units_m = re.search(r"已發行受益權單位總數\s*([\d,]+)", text)
        date_m = re.search(r"公告日期\s*[:：]?\s*(\d{4}/\d{2}/\d{2})", text)
        nav_date = date_m.group(1).replace("/", "-") if date_m else None

        return _check_aum_nav_units_consistency(
            _fund_assets_or_none(
                aum=aum_m.group(1) if aum_m else None,
                nav=nav_m.group(1) if nav_m else None,
                units=units_m.group(1) if units_m else None,
                nav_date=nav_date,
            )
        )
    except Exception as exc:
        logger.error("[Yuanta] %s 資產摘要抓取失敗（不影響持股）：%s", fund_code, exc)
        return None


# ── 台新 (00987A) ─────────────────────────────────────────────────────────────


def _fetch_taishin(etf_code: str) -> list[dict[str, Any]]:
    """Fetch 00987A holdings via Taishin official HTML page (server-side rendered)."""
    from bs4 import BeautifulSoup

    try:
        url = f"https://www.tsit.com.tw/ETF/Home/ETFSeriesDetail/{etf_code}"
        raw = _get(url, verify_ssl=False)
        html = raw.decode("utf-8", errors="replace")
        soup = BeautifulSoup(html, "lxml")

        # Extract data date from <input id="PUB_DATE" value="YYYY-MM-DD">
        pub_date_input = soup.find("input", id="PUB_DATE")
        data_date = (pub_date_input.get("value") or "") if pub_date_input else ""

        # Locate holdings table: must have <th> cells containing both "代號" and "股數"
        # (excludes the fund-info summary table which only has "代號" / "中文名稱" etc.)
        table = next(
            (
                t
                for t in soup.find_all("table")
                if any("代號" in th.get_text(strip=True) for th in t.find_all("th"))
                and any("股數" in th.get_text(strip=True) for th in t.find_all("th"))
            ),
            None,
        )
        if table is None:
            logger.warning("[Taishin] %s 找不到持股表格（含代號+股數欄位）", etf_code)
            return []

        # Codes may have exchange suffix like "6274 TT" — strip trailing uppercase alpha
        code_re = re.compile(r"^\d{4,6}$")
        exchange_sfx_re = re.compile(r"\s+[A-Z]{2,}$")
        holdings: list[dict[str, Any]] = []
        for tr in table.find_all("tr"):
            tds = tr.find_all("td")
            if len(tds) < 4:
                continue
            raw_code = tds[0].get_text(strip=True)
            code = exchange_sfx_re.sub("", raw_code).strip()
            if not code_re.match(code):
                continue
            name = tds[1].get_text(strip=True)
            shares_raw = tds[2].get_text(strip=True).replace(",", "")
            weight_raw = tds[3].get_text(strip=True).replace("%", "").strip()
            try:
                shares = int(shares_raw)
            except ValueError:
                shares = 0
            w = _to_num(weight_raw)
            if w is None:
                continue
            holdings.append(
                {"code": code, "name": name, "shares": shares, "weight_pct": w}
            )

        logger.info(
            "[Taishin] %s 取得 %d 筆持股（資料日期：%s）",
            etf_code,
            len(holdings),
            data_date,
        )
        return holdings
    except Exception as exc:
        logger.error("[Taishin] %s 抓取失敗：%s", etf_code, exc)
        return []


def _fetch_taishin_assets(etf_code: str, date_ymd: str | None) -> dict[str, Any] | None:
    """台新 ETF 資產摘要：AJAX 端點 PartialHistoryFundNav（頁面「歷史淨值」分頁資料來源）。

    回傳最新一筆 {淨值, nav_date}；台新未揭露 aum/units，故 aum/units 恆 None。
    失敗回 None，不影響持股。
    """
    from bs4 import BeautifulSoup

    try:
        eday = _ymd_to_dash(date_ymd) if date_ymd else _last_weekday_dash()
        from datetime import datetime, timedelta

        sday = (datetime.fromisoformat(eday) - timedelta(days=10)).date().isoformat()
        url = "https://www.tsit.com.tw/ETF/Home/PartialHistoryFundNav"
        payload = f"equal=0&id={etf_code}&sday={sday}&eday={eday}"
        req = urllib.request.Request(
            url,
            data=payload.encode("utf-8"),
            headers={
                "User-Agent": UA,
                "Content-Type": "application/x-www-form-urlencoded",
            },
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=30, context=_SSL_NO_VERIFY) as resp:
            raw = resp.read()
        html = raw.decode("utf-8", errors="replace")
        soup = BeautifulSoup(html, "lxml")
        table = soup.find("table")
        if table is None:
            return None
        rows = table.find_all("tr")
        for tr in rows:
            tds = tr.find_all("td")
            if len(tds) < 2:
                continue
            date_txt = tds[0].get_text(strip=True)
            nav_txt = tds[1].get_text(strip=True)
            if not re.match(r"^\d{4}-\d{2}-\d{2}$", date_txt):
                continue
            return _fund_assets_or_none(nav=nav_txt, nav_date=date_txt)
        return None
    except Exception as exc:
        logger.error("[Taishin] %s 資產摘要抓取失敗（不影響持股）：%s", etf_code, exc)
        return None


# ── 第一金 (00994A) ────────────────────────────────────────────────────────────


def _fetch_first_financial(fund_id: str) -> list[dict[str, Any]]:
    """Fetch 00994A holdings via First Financial REST API (double-decode JSON response)."""
    url = "https://www.fsitc.com.tw/WebAPI.aspx/Get_hd"
    payload: dict[str, Any] = {"pStrFundID": fund_id, "pStrDate": ""}
    raw = _post_json(url, payload, verify_ssl=False)
    outer: dict[str, Any] = json.loads(raw.decode("utf-8"))
    inner_str = outer.get("d") or "[]"
    items: list[dict[str, Any]] = json.loads(inner_str)

    # Live response uses single-letter keys: A=code, B=name, C=weight%, D=shares
    holdings: list[dict[str, Any]] = []
    for item in items:
        code = str(item.get("A") or item.get("stockCode") or "").strip()
        name = str(item.get("B") or item.get("stockName") or "").strip()
        weight = _to_num(item.get("C") or item.get("weight"))
        shares = _to_num(
            (item.get("D") or "").replace(",", "")
            if isinstance(item.get("D"), str)
            else item.get("D")
        )
        if code and weight is not None:
            holdings.append(
                {"code": code, "name": name, "shares": shares, "weight_pct": weight}
            )
    return holdings


# ── 中信 (00995A) ──────────────────────────────────────────────────────────────


def _parse_ctbc_fund_assets(data: dict[str, Any]) -> dict[str, Any] | None:
    """中信 ETFHoldingWeight 回應 Data.FundAssets[0] 資產摘要解析。

    中文 key（基金淨資產/基金每單位淨值/基金在外流通單位數）與 NAV_DT
    格式（ISO 含 T，截前段）均為 2026-07-12 實測回應。失敗回 None，不影響持股。
    """
    try:
        fund_assets_list: list[dict[str, Any]] = data.get("FundAssets") or []
        if not fund_assets_list:
            return None
        fa = fund_assets_list[0]
        nav_dt = str(fa.get("NAV_DT") or "")
        nav_date = nav_dt.split("T")[0] if nav_dt else None
        return _check_aum_nav_units_consistency(
            _fund_assets_or_none(
                aum=fa.get("基金淨資產"),
                nav=fa.get("基金每單位淨值"),
                units=fa.get("基金在外流通單位數"),
                nav_date=nav_date,
            )
        )
    except Exception as exc:
        logger.error("[CTBC] 資產摘要解析失敗（不影響持股）：%s", exc)
        return None


def _fetch_ctbc(
    fid: str, date_ymd: str | None = None
) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
    """Fetch 00995A holdings + 資產摘要 via CTBC two-step authenticated REST API.

    Step 1: POST AuthToken to obtain a short-lived token (never cached).
    Step 2: POST ETFHoldingWeight — StartDate 必須帶 dash 格式日期，
        空字串會使 API 回 ResultCode 1 且無資料（2026-07-12 實測）。
    持股取 Data.FundAssetsDetail 中 Code == "STOCK" section 的 Data 清單，
    欄位 code_/name_/weights_/qty_（qty_ 含千分位逗號）；資產摘要取同回應
    的 Data.FundAssets[0]，零額外請求。
    """
    # Step 1 — get auth token。
    # token 必須「同時」帶於 URL query string 與 JSON body：僅放 body 會被 API
    # 回 ResultCode 1「Token 無效或過期」；seed 為含 .tw 的完整網域（2026-07-12 實測）。
    seed = "www.ctbcinvestments.com.tw"
    auth_raw = _post_json(
        "https://www.ctbcinvestments.com.tw/API/home/AuthToken?token="
        + urllib.parse.quote(seed),
        {"token": seed},
        verify_ssl=False,
    )
    auth_js: dict[str, Any] = json.loads(auth_raw.decode("utf-8"))
    token: str | None = (auth_js.get("Data") or {}).get("token")
    if not token:
        raise RuntimeError(f"中信 AuthToken 未取得 token，回應：{auth_js}")

    # Step 2 — fetch holdings + fund assets
    start_date = _ymd_to_dash(date_ymd) if date_ymd else _last_weekday_dash()
    holdings_raw = _post_json(
        "https://www.ctbcinvestments.com.tw/API/etf/ETFHoldingWeight?token="
        + urllib.parse.quote(token),
        {"token": token, "FID": fid, "StartDate": start_date},
        verify_ssl=False,
    )
    holdings_js: dict[str, Any] = json.loads(holdings_raw.decode("utf-8"))
    if holdings_js.get("ResultCode") != 0:
        logger.warning(
            "[CTBC] %s ETFHoldingWeight ResultCode=%s（StartDate=%s），回空持股走 fallback",
            fid,
            holdings_js.get("ResultCode"),
            start_date,
        )
        return [], None
    data: dict[str, Any] = holdings_js.get("Data") or {}

    sections: list[dict[str, Any]] = data.get("FundAssetsDetail") or []
    stock_section = next((s for s in sections if s.get("Code") == "STOCK"), None)
    holdings: list[dict[str, Any]] = []
    for item in (stock_section or {}).get("Data") or []:
        code = str(item.get("code_") or "").strip()
        name = str(item.get("name_") or "").strip()
        shares = _to_num(item.get("qty_"))
        weight = _to_num(item.get("weights_"))
        if code and weight is not None:
            holdings.append(
                {"code": code, "name": name, "shares": shares, "weight_pct": weight}
            )
    if not holdings:
        logger.warning("[CTBC] %s 回應無 STOCK section 或無有效持股", fid)

    return holdings, _parse_ctbc_fund_assets(data)


# ── 兆豐 (00996A) ──────────────────────────────────────────────────────────────


def _parse_mega_assets(html: str) -> dict[str, Any] | None:
    """兆豐 `#asset_div`「基金資產」區塊解析。

    頁面含一段結構相同的 HTML 註解（範例假資料），必須先剝除註解再解析，
    否則會誤抓到假資料。失敗回 None，不影響持股。
    """
    from bs4 import BeautifulSoup, Comment

    try:
        soup = BeautifulSoup(html, "lxml")
        # 剝除所有 HTML 註解（含註解內可能藏的假資料 asset_div）
        for c in soup.find_all(string=lambda s: isinstance(s, Comment)):
            c.extract()
        asset_div = soup.find(id="asset_div")
        if asset_div is None:
            return None
        text = asset_div.get_text(" ", strip=True)
        aum_m = re.search(r"淨資產價值\s*([\d,]+)", text)
        units_m = re.search(r"在外流通單位數\s*([\d,]+)", text)
        nav_m = re.search(r"每單位淨值\s*([\d,.]+)", text)
        date_m = re.search(r"(\d{4})/(\d{2})/(\d{2})", text)
        nav_date = (
            f"{date_m.group(1)}-{date_m.group(2)}-{date_m.group(3)}" if date_m else None
        )
        return _check_aum_nav_units_consistency(
            _fund_assets_or_none(
                aum=aum_m.group(1) if aum_m else None,
                nav=nav_m.group(1) if nav_m else None,
                units=units_m.group(1) if units_m else None,
                nav_date=nav_date,
            )
        )
    except Exception as exc:
        logger.error("[MEGA] 資產摘要解析失敗（不影響持股）：%s", exc)
        return None


def _fetch_mega(
    fund_id: str | None,
) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
    """Fetch 兆豐 ETF holdings via Mega Funds official HTML page.

    fund_id must be confirmed by browsing https://www.megafunds.com.tw/MEGA/etf/.
    If fund_id is None or empty, returns [] and logs a warning so multi_etf_step
    can fall back to pocket_scraper.
    """
    from bs4 import BeautifulSoup

    if not fund_id:
        logger.warning("[MEGA] fund_id not configured for 00986A")
        return [], None
    try:
        url = f"https://www.megafunds.com.tw/MEGA/etf/etf_product.aspx?id={fund_id}"
        raw = _get(url)
        html = raw.decode("utf-8", errors="replace")
        soup = BeautifulSoup(html, "lxml")

        # Extract data date from first YYYY/MM/DD occurrence in page text
        date_match = re.search(r"(\d{4})/(\d{2})/(\d{2})", html)
        data_date = (
            f"{date_match.group(1)}-{date_match.group(2)}-{date_match.group(3)}"
            if date_match
            else ""
        )

        assets = _parse_mega_assets(html)

        content_div = soup.find("div", id="fund_content_list_1")
        if content_div is None:
            logger.warning("[MEGA] 找不到 fund_content_list_1 容器")
            return [], assets

        code_re = re.compile(r"^\d{4,6}$")
        holdings: list[dict[str, Any]] = []
        for fund_info in content_div.find_all("div", class_="fund-info"):
            cols = fund_info.find_all("div", class_="fund-content")
            if len(cols) < 4:
                continue
            code = cols[0].get_text(strip=True)
            if not code_re.match(code):
                continue
            name = cols[1].get_text(strip=True)
            shares_raw = cols[2].get_text(strip=True).replace(",", "")
            weight_raw = cols[3].get_text(strip=True).replace("%", "").strip()
            try:
                shares = int(shares_raw)
            except ValueError:
                shares = 0
            w = _to_num(weight_raw)
            if w is None:
                continue
            holdings.append(
                {"code": code, "name": name, "shares": shares, "weight_pct": w}
            )

        logger.info("[MEGA] 取得 %d 筆持股（資料日期：%s）", len(holdings), data_date)
        return holdings, assets
    except Exception as exc:
        logger.error("[MEGA] 抓取失敗：%s", exc)
        return [], None


# ── 摩根 JPM (00401A, 00989A) ─────────────────────────────────────────────────


def _fetch_jpm(
    xlsx_url: str,
) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
    """Fetch JPM ETF holdings from a fixed XLSX URL.

    The XLSX uses a two-section layout:
    - Summary section: header row with 'Fund Ticker' + 'Estimated Total Market Value'
    - Detail section: header row with 'Constituent Ticker', then 'D' rows
    weight = Market Value Base / Estimated Total Market Value * 100

    資產摘要另從 summary 'S' row 讀 'Estimated NAV per Share' / 'Outstanding Shares'。
    'Valuation Date' 欄位經人工複核發現可能是未來結算日（非 NAV 揭露日），
    不可靠，故 nav_date 恆留 None（DB 寫入日期另由 pipeline ctx.date_str 決定）。
    """
    import openpyxl

    try:
        req = urllib.request.Request(
            xlsx_url,
            headers={
                "User-Agent": UA,
                "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
                "Referer": "https://am.jpmorgan.com/tw/zh/asset-management/per/",
            },
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = resp.read()
    except Exception as exc:
        logger.error("[JPM] XLSX 下載失敗：%s", exc)
        return [], None

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        if ws is None:
            return [], None
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        logger.error("[JPM] XLSX 解析失敗：%s", exc)
        return [], None

    def cells(row: tuple[Any, ...]) -> list[str]:
        return [str(c).strip() if c is not None else "" for c in row]

    # Locate header rows
    summary_hdr_idx = -1
    detail_hdr_idx = -1
    for i, row in enumerate(rows):
        if not row:
            continue
        cs = cells(row)
        if "Record Type" in cs and "Fund Ticker" in cs and summary_hdr_idx == -1:
            summary_hdr_idx = i
        if "Record Type" in cs and "Constituent Ticker" in cs and detail_hdr_idx == -1:
            detail_hdr_idx = i

    if summary_hdr_idx == -1 or detail_hdr_idx == -1:
        logger.warning("[JPM] XLSX 結構未識別（找不到 header rows），URL=%s", xlsx_url)
        return [], None

    # Extract Estimated Total Market Value / NAV per Share / Outstanding Shares
    # from summary "S" row
    s_headers = cells(rows[summary_hdr_idx])
    etmv_col = next(
        (i for i, h in enumerate(s_headers) if "Estimated Total Market Value" in h), -1
    )
    aum_col = next((i for i, h in enumerate(s_headers) if h == "Estimated NAV"), -1)
    nav_col = next(
        (i for i, h in enumerate(s_headers) if "Estimated NAV per Share" in h), -1
    )
    shares_out_col = next(
        (i for i, h in enumerate(s_headers) if "Outstanding Shares" in h), -1
    )
    total_market_value = 0.0
    assets: dict[str, Any] | None = None
    if etmv_col >= 0:
        for row in rows[summary_hdr_idx + 1 : detail_hdr_idx]:
            if not row or all(c is None for c in row):
                continue
            v = _to_num(row[etmv_col]) if etmv_col < len(row) else None
            if v is not None and v > 0:
                total_market_value = v
            aum_val = (
                _to_num(row[aum_col]) if aum_col >= 0 and aum_col < len(row) else None
            )
            nav_val = (
                _to_num(row[nav_col]) if nav_col >= 0 and nav_col < len(row) else None
            )
            units_val = (
                _to_num(row[shares_out_col])
                if shares_out_col >= 0 and shares_out_col < len(row)
                else None
            )
            if aum_val is not None or nav_val is not None or units_val is not None:
                # 'Valuation Date' 欄位經人工複核發現可能是未來結算日，不可靠，
                # nav_date 恆留 None（見函式 docstring）。
                assets = _check_aum_nav_units_consistency(
                    _fund_assets_or_none(
                        aum=aum_val, nav=nav_val, units=units_val, nav_date=None
                    )
                )
            if v is not None and v > 0:
                break

    # Parse detail rows
    d_headers = cells(rows[detail_hdr_idx])

    def col(name: str) -> int:
        return next((i for i, h in enumerate(d_headers) if name in h), -1)

    rt_col = col("Record Type")
    ticker_col = col("Constituent Ticker")
    name_col = col("Constituent Description")
    shares_col = col("Shares or PAR Amount")
    mv_col = col("Market Value Base")

    if ticker_col == -1 or mv_col == -1:
        logger.warning(
            "[JPM] XLSX 缺少必要欄位（Constituent Ticker / Market Value Base）"
        )
        return [], assets

    holdings: list[dict[str, Any]] = []
    for row in rows[detail_hdr_idx + 1 :]:
        if not row or all(c is None for c in row):
            if holdings:
                break
            continue
        if rt_col >= 0 and len(row) > rt_col:
            rt_val = str(row[rt_col]).strip() if row[rt_col] is not None else ""
            if rt_val != "D":
                continue
        code = (
            str(row[ticker_col]).strip()
            if ticker_col < len(row) and row[ticker_col] is not None
            else ""
        )
        if not code:
            continue
        name = (
            str(row[name_col]).strip()
            if name_col >= 0 and name_col < len(row) and row[name_col] is not None
            else ""
        )
        shares_raw = (
            _to_num(row[shares_col])
            if shares_col >= 0 and shares_col < len(row)
            else None
        )
        mv_raw = (
            _to_num(row[mv_col])
            if mv_col < len(row) and row[mv_col] is not None
            else None
        )
        if mv_raw is None:
            continue
        weight_pct = (
            (mv_raw / total_market_value * 100) if total_market_value > 0 else 0.0
        )
        holdings.append(
            {
                "code": code,
                "name": name,
                "shares": int(shares_raw) if shares_raw is not None else 0,
                "weight_pct": weight_pct,
            }
        )

    logger.info("[JPM] %s 取得 %d 筆持股", xlsx_url.split("/")[-1], len(holdings))
    return holdings, assets


# ── 國泰 Cathay (00400A) ───────────────────────────────────────────────────────


def _fetch_cathay(fund_code: str) -> list[dict[str, Any]]:
    """Fetch 國泰 ETF holdings via REST GET endpoint.

    shares is always 0 — 國泰官方不揭露股數。
    """
    url = f"https://cwapi.cathaysite.com.tw/api/ETF/GetIndexStockWeights?fundCode={fund_code}"
    try:
        raw = _get(url)
        js: dict[str, Any] = json.loads(raw.decode("utf-8"))
    except Exception as exc:
        logger.error("[Cathay] %s 抓取失敗：%s", fund_code, exc)
        return []

    stock_weights: list[dict[str, Any]] = (js.get("result") or {}).get(
        "stockWeights"
    ) or []
    if not stock_weights:
        logger.warning("[Cathay] %s 回傳空 stockWeights", fund_code)
        return []

    holdings: list[dict[str, Any]] = []
    for sw in stock_weights:
        code = str(sw.get("stockCode") or "").strip()
        name = str(sw.get("stockName") or "").strip()
        weight_pct = _to_num(sw.get("weights"))
        if code and weight_pct is not None:
            holdings.append(
                {"code": code, "name": name, "shares": 0, "weight_pct": weight_pct}
            )

    logger.info("[Cathay] %s 取得 %d 筆持股", fund_code, len(holdings))
    return holdings


def _fetch_cathay_assets(fund_code: str) -> dict[str, Any] | None:
    """國泰 GetETFAssets 資產摘要（持股端點不含摘要，允許的一次補充請求）。

    命名陷阱（2026-07-12 實測）：`fundNav` 是基金「總淨資產」（aum，~257 億），
    每單位淨值是 `fundPerNav`（nav，~14.2），不可對調。
    任何失敗（HTTP 錯誤、success != true、欄位缺漏）只 log 回 None，不影響持股。
    """
    url = f"https://cwapi.cathaysite.com.tw/api/ETF/GetETFAssets?fundCode={fund_code}"
    try:
        raw = _get(url)
        js: dict[str, Any] = json.loads(raw.decode("utf-8"))
        if not js.get("success"):
            logger.error(
                "[Cathay] %s GetETFAssets success != true（不影響持股）", fund_code
            )
            return None
        result: dict[str, Any] = js.get("result") or {}
        pre_date = str(result.get("preDate") or "").strip().replace("/", "-")
        return _check_aum_nav_units_consistency(
            _fund_assets_or_none(
                aum=result.get("fundNav"),
                nav=result.get("fundPerNav"),
                units=result.get("fundOutstandingShares"),
                nav_date=pre_date or None,
            )
        )
    except Exception as exc:
        logger.error(
            "[Cathay] %s GetETFAssets 抓取失敗（不影響持股）：%s", fund_code, exc
        )
        return None


# ── 聯博 AllianceBernstein (00984D) ───────────────────────────────────────────


def _fetch_alliance_bernstein(fund_code: str) -> list[dict[str, Any]]:
    """Fetch 00984D holdings via AllianceBernstein REST API.

    Parses domesticHoldings[].holdings[].
    Accepts any non-empty code (bond ISINs, not just 4-digit TW codes).
    """
    url = f"https://www.abfunds.com.tw/api/fundholding/{fund_code}"
    try:
        raw = _get(url)
        js: dict[str, Any] = json.loads(raw.decode("utf-8"))
    except Exception as exc:
        logger.error("[AllianceBernstein] %s 抓取失敗：%s", fund_code, exc)
        return []

    domestic_holdings: list[dict[str, Any]] = js.get("domesticHoldings") or []
    if not domestic_holdings:
        logger.warning("[AllianceBernstein] %s 回傳空 domesticHoldings", fund_code)
        return []

    holdings: list[dict[str, Any]] = []
    for section in domestic_holdings:
        for item in section.get("holdings") or []:
            code = str(item.get("stockCode") or item.get("code") or "").strip()
            if not code:
                continue
            name = str(item.get("stockName") or item.get("name") or "").strip()
            shares_val = item.get("shares") or item.get("quantity")
            shares_num = _to_num(shares_val)
            weight = _to_num(item.get("weight") or item.get("weightPct"))
            if weight is not None:
                holdings.append(
                    {
                        "code": code,
                        "name": name,
                        "shares": int(shares_num) if shares_num is not None else 0,
                        "weight_pct": weight,
                    }
                )

    logger.info("[AllianceBernstein] %s 取得 %d 筆持股", fund_code, len(holdings))
    return holdings


# ── 富邦 Fubon (00982D, 00983D) ────────────────────────────────────────────────


def _parse_fubon_assets(soup: Any) -> dict[str, Any] | None:
    """富邦 `<div class="fund_box p2">` 內 `<ul><li>` 資產摘要解析，與持股明細段落無關。"""
    try:
        box = soup.find("div", class_="fund_box")
        text = box.get_text(" ", strip=True) if box else soup.get_text(" ", strip=True)
        aum_m = re.search(r"基金淨資產\(新台幣\)\s*([\d,]+)", text)
        units_m = re.search(r"基金在外流通單位數\(單位\)\s*([\d,]+)", text)
        nav_m = re.search(r"基金每單位淨值\(新台幣\)\s*([\d,.]+)", text)
        # 資料日期緊接在 fund_box 之後（非 box 內），box 內找不到就搜全頁
        date_m = re.search(r"資料日期\s*[:：]\s*(\d{4}/\d{2}/\d{2})", text)
        if date_m is None and box is not None:
            date_m = re.search(
                r"資料日期\s*[:：]\s*(\d{4}/\d{2}/\d{2})",
                soup.get_text(" ", strip=True),
            )
        nav_date = date_m.group(1).replace("/", "-") if date_m else None
        return _check_aum_nav_units_consistency(
            _fund_assets_or_none(
                aum=aum_m.group(1) if aum_m else None,
                nav=nav_m.group(1) if nav_m else None,
                units=units_m.group(1) if units_m else None,
                nav_date=nav_date,
            )
        )
    except Exception as exc:
        logger.error("[Fubon] 資產摘要解析失敗（不影響持股）：%s", exc)
        return None


def _fetch_fubon(
    fund_code: str,
) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
    """Fetch 富邦 ETF holdings via official HTML page.

    Locates <h6> containing "持股明細" and reads subsequent <tbody>.
    Bond ETFs (00982D, 00983D) have no equity section → returns [] gracefully.
    NAV 資產摘要位於同頁 <ul><li> 區塊，與持股明細段落無關，不受無持股影響。
    """
    from bs4 import BeautifulSoup

    try:
        url = f"https://websys.fsit.com.tw/FubonETF/Trade/Assets.aspx?lan=TW&stkId={fund_code}"
        # 富邦憑證鏈缺 Subject Key Identifier，同 taishin/ctbc 需停用驗證
        raw = _get(url, verify_ssl=False)
        html = raw.decode("utf-8", errors="replace")
        soup = BeautifulSoup(html, "lxml")
    except Exception as exc:
        logger.error("[Fubon] %s 抓取失敗：%s", fund_code, exc)
        return [], None

    assets = _parse_fubon_assets(soup)

    h6 = next(
        (tag for tag in soup.find_all("h6") if "持股明細" in tag.get_text(strip=True)),
        None,
    )
    if h6 is None:
        logger.warning("[Fubon] %s 找不到 <h6>持股明細（可能為債券型 ETF）", fund_code)
        return [], assets

    tbody = h6.find_next("tbody")
    if tbody is None:
        logger.warning("[Fubon] %s 找不到持股 <tbody>", fund_code)
        return [], assets

    code_re = re.compile(r"^\d{4,6}$")
    holdings: list[dict[str, Any]] = []
    for tr in tbody.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 4:
            continue
        code = tds[0].get_text(strip=True)
        if not code_re.match(code):
            continue
        name = tds[1].get_text(strip=True)
        shares_raw = tds[2].get_text(strip=True).replace(",", "")
        weight_raw = tds[3].get_text(strip=True).replace("%", "").strip()
        shares_num = _to_num(shares_raw)
        shares = int(shares_num) if shares_num is not None else 0
        w = _to_num(weight_raw)
        if w is None:
            continue
        holdings.append({"code": code, "name": name, "shares": shares, "weight_pct": w})

    logger.info("[Fubon] %s 取得 %d 筆持股", fund_code, len(holdings))
    return holdings, assets


# ── 中信 HTML (00983A) ─────────────────────────────────────────────────────────


def _fetch_ctbc_html(
    etf_code: str,
) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
    """Fetch 中信 ARK ETF holdings via legacy ASP.NET HTML page.

    Uses pcd.aspx (not the auth-token REST API used for 00995A).
    SSL verification disabled for consistency with other CTBC calls.
    資產摘要：Label_AUM01(日期)/02(淨資產)/03(流通單位)/04(每單位淨值) 既有頁面 selector。
    """
    from bs4 import BeautifulSoup

    try:
        url = f"https://www.ctbcinvestments.com.tw/CTWEB/Content/ETF/pcd.aspx?ETF_ID={etf_code}"
        raw = _get(url, verify_ssl=False)
        html = raw.decode("utf-8", errors="replace")
        soup = BeautifulSoup(html, "lxml")
    except Exception as exc:
        logger.error("[CTBC_HTML] %s 抓取失敗：%s", etf_code, exc)
        return [], None

    # Log disclosure date (Label_AUM01)
    aum_label = soup.find(id="Label_AUM01")
    data_date = aum_label.get_text(strip=True) if aum_label else ""
    logger.info("[CTBC_HTML] %s 揭露日期：%s", etf_code, data_date)

    assets: dict[str, Any] | None = None
    try:
        aum02 = soup.find(id="Label_AUM02")
        aum03 = soup.find(id="Label_AUM03")
        aum04 = soup.find(id="Label_AUM04")
        assets = _check_aum_nav_units_consistency(
            _fund_assets_or_none(
                aum=aum02.get_text(strip=True) if aum02 else None,
                units=aum03.get_text(strip=True) if aum03 else None,
                nav=aum04.get_text(strip=True) if aum04 else None,
                nav_date=data_date or None,
            )
        )
    except Exception as exc:
        logger.error("[CTBC_HTML] %s 資產摘要解析失敗（不影響持股）：%s", etf_code, exc)

    # 支援台股 4 位數代號（2330）與美股代號（TSLA US, AMD US）
    code_re = re.compile(r"^\d{4,6}$|^[A-Z0-9]{2,10}(\s+[A-Z]{2,3})?$")
    holdings: list[dict[str, Any]] = []
    for tr in soup.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 4:
            continue
        code = tds[0].get_text(strip=True)
        if not code_re.match(code):
            continue
        name = tds[1].get_text(strip=True)
        shares_raw = tds[2].get_text(strip=True).replace(",", "")
        weight_raw = tds[3].get_text(strip=True).replace("%", "").strip()
        shares_num = _to_num(shares_raw)
        shares = int(shares_num) if shares_num is not None else 0
        w = _to_num(weight_raw)
        if w is None:
            continue
        holdings.append({"code": code, "name": name, "shares": shares, "weight_pct": w})

    logger.info("[CTBC_HTML] %s 取得 %d 筆持股", etf_code, len(holdings))
    return holdings, assets


# ── Date utils ────────────────────────────────────────────────────────────────


def _ymd_to_dash(ymd: str) -> str:
    if "-" in ymd:
        return ymd
    return f"{ymd[:4]}-{ymd[4:6]}-{ymd[6:8]}"


def _last_trading_date():
    from datetime import datetime, timedelta, timezone

    tw_now = datetime.now(timezone(timedelta(hours=8)))
    if tw_now.hour < 15:
        tw_now -= timedelta(days=1)
    d = tw_now.date()
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d


def _last_weekday_ymd() -> str:
    return _last_trading_date().strftime("%Y%m%d")


def _last_weekday_dash() -> str:
    return _last_trading_date().isoformat()


def _to_ymd(date_str: str | None) -> str | None:
    """YYYY-MM-DD → YYYYMMDD（已是 YYYYMMDD 則原樣回傳）"""
    if not date_str:
        return None
    return date_str.replace("-", "")


# ── Public interface ──────────────────────────────────────────────────────────


def fetch_holdings(etf_code: str, date_str: str | None = None) -> pd.DataFrame:
    """從投信官網 API 取得持股，回傳標準化 DataFrame。

    Args:
        etf_code: ETF 代號（如 "00981A"）。
        date_str: 資料日期，YYYY-MM-DD 或 YYYYMMDD；None 表示最新交易日。

    Returns:
        DataFrame，欄位：code(str), name(str), weight(float,%), shares(int)。
        ETF 代號不在 CATALOG 或發生任何錯誤時回傳空 DataFrame，不拋例外。
    """
    if not is_supported(etf_code):
        logger.warning("[OFFICIAL_API] %s 不在 CATALOG，回傳空 DataFrame", etf_code)
        return pd.DataFrame(columns=["code", "name", "weight", "shares"])

    cat = CATALOG[etf_code]
    issuer = cat["issuer"]
    fund_code = cat["fund_code"]
    date_ymd = _to_ymd(date_str)

    try:
        raw_holdings, fund_assets = _dispatch(issuer, fund_code, date_ymd, cat)
        if not raw_holdings:
            logger.warning("[OFFICIAL_API] %s 回傳空持股", etf_code)
            empty = pd.DataFrame(columns=["code", "name", "weight", "shares"])
            # 債券型 ETF（如富邦 00982D/00983D）無股票持股但仍有資產摘要，
            # 一併附上供 MultiEtfStep 取用，不因空持股而遺失 NAV。
            if fund_assets:
                empty.attrs["fund_assets"] = fund_assets
            return empty

        rows = [
            {
                "code": h["code"],
                "name": h["name"],
                "weight": float(h["weight_pct"] or 0),
                "shares": int(h["shares"] or 0),
            }
            for h in raw_holdings
            if h.get("code") and h.get("weight_pct") is not None
        ]
        df = pd.DataFrame(rows, columns=["code", "name", "weight", "shares"])
        # 基金資產摘要附於 df.attrs（同 moneydj_scraper 的 data_date 慣例），
        # 供 MultiEtfStep 取出寫入 ctx.etf_fund_assets。無摘要則不附。
        if fund_assets:
            df.attrs["fund_assets"] = fund_assets
        logger.info("[OFFICIAL_API] %s 取得 %d 筆持股", etf_code, len(df))
        return df

    except Exception as e:
        logger.error("[OFFICIAL_API] %s 抓取失敗：%s", etf_code, e)
        return pd.DataFrame(columns=["code", "name", "weight", "shares"])


def _dispatch(
    issuer: str,
    fund_code: str,
    date_ymd: str | None,
    cat_entry: dict[str, Any] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
    """回傳 (holdings, fund_assets)。

    holdings 為持股 list；fund_assets 為 {aum, nav, units, nav_date}，
    來源未揭露基金資產摘要時為 None（XLSX / HTML / 其他僅持股的來源）。
    """
    # ── 帶基金資產摘要的 JSON 來源 ──
    if issuer == "nomura":
        return _fetch_nomura(fund_code, date_ymd)
    if issuer == "allianz":
        return _fetch_allianz(fund_code, date_ymd)
    if issuer == "capital":
        return _fetch_capital(fund_code, date_ymd)

    # ── 現已補資產摘要（XLSX / HTML 同頁擴充解析）──
    if issuer == "uni":
        if date_ymd:
            logger.warning("[OFFICIAL_API] 統一 ezmoney 不支援指定日期，改抓最新")
        return _fetch_uni(fund_code)
    if issuer == "fhtrust":
        ymd = date_ymd or _last_weekday_ymd()
        return _fetch_fhtrust(fund_code, ymd)
    if issuer == "yuanta":
        from ETF.scrapers import yuanta_scraper

        df = yuanta_scraper.fetch_holdings(fund_code)
        records = df.rename(columns={"weight": "weight_pct"}).to_dict("records")
        holdings = cast(list[dict[str, Any]], records) if not df.empty else []
        assets = _fetch_yuanta_assets(fund_code)
        return holdings, assets
    if issuer == "taishin":
        holdings = _fetch_taishin(fund_code)
        assets = _fetch_taishin_assets(fund_code, date_ymd)
        return holdings, assets
    if issuer == "mega":
        fund_id: str | None = (cat_entry or {}).get("fund_id")
        return _fetch_mega(fund_id)
    if issuer == "jpm":
        xlsx_url: str = (cat_entry or {}).get("xlsx_url", "")
        return _fetch_jpm(xlsx_url)
    if issuer == "ctbc_html":
        return _fetch_ctbc_html(fund_code)
    if issuer == "fubon":
        return _fetch_fubon(fund_code)
    if issuer == "ctbc":
        return _fetch_ctbc(fund_code, date_ymd)
    if issuer == "cathay":
        return _fetch_cathay(fund_code), _fetch_cathay_assets(fund_code)

    # ── 僅持股、無資產摘要的來源 ──
    if issuer == "first_financial":
        return _fetch_first_financial(fund_code), None
    if issuer == "alliance_bernstein":
        return _fetch_alliance_bernstein(fund_code), None
    raise RuntimeError(f"未知 issuer: {issuer}")
