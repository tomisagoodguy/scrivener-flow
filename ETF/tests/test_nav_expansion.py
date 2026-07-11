"""Unit tests for NAV/AUM/units coverage expansion (nav-expansion task).

Covers newly added `fund_assets` extraction for: uni, fhtrust, yuanta, taishin,
mega, ctbc_html, fubon, jpm. All HTTP calls are mocked; no network access.
"""

from __future__ import annotations

import io
import unittest
from unittest.mock import MagicMock, patch


# ── Helpers ────────────────────────────────────────────────────────────────────


def _make_uni_xlsx() -> bytes:
    """統一 ezmoney XLSX：資產摘要 3 列（民國年日期）+ 持股表頭。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["淨資產：NTD 274,753,577,961"])
    ws.append(["流通在外單位數：9,199,209,000"])
    ws.append(["每單位淨值：NTD 29.87"])
    ws.append(["資料日期：115/07/09"])
    ws.append([])
    ws.append(["代號", "名稱", "股數", "權重"])
    ws.append(["2330", "台積電", 1000, 10.5])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_fhtrust_xlsx() -> bytes:
    """復華 XLSX：資產摘要（西元年日期）+ 持股表頭。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["基金資產淨值：75,346,169,965"])
    ws.append(["基金在外流通單位數：3,882,416,000"])
    ws.append(["基金每單位淨值", 19.41])
    ws.append(["日期: 2026/07/09"])
    ws.append([])
    ws.append(["代號", "名稱", "股數", "比例"])
    ws.append(["2330", "台積電", 1000, 10.5])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_mega_html(with_comment_decoy: bool = True) -> str:
    decoy = (
        '<!-- <div id="asset_div">淨資產價值 999,999,999 在外流通單位數 '
        "999,999,999 每單位淨值 99.99 2000/01/01</div> -->"
        if with_comment_decoy
        else ""
    )
    return f"""<html><body>
    {decoy}
    <div id="asset_div">
      淨資產價值 4,848,922,587
      在外流通單位數 321,899,000
      每單位淨值 15.06
      資料來源：兆豐投信，2026/07/09
    </div>
    <div id="fund_content_list_1"></div>
    </body></html>"""


def _make_ctbc_html_with_assets() -> bytes:
    return """<html><body>
    <span id="Label_AUM01">2026/07/08</span>
    <span id="Label_AUM02">1,297,629,321</span>
    <span id="Label_AUM03">105,686,000</span>
    <span id="Label_AUM04">12.28</span>
    <table>
      <tr><th>代號</th><th>名稱</th><th>股數</th><th>比重(%)</th></tr>
      <tr><td>2330</td><td>台積電</td><td>10,000</td><td>9.78%</td></tr>
    </table>
    </body></html>""".encode("utf-8")


def _make_fubon_html() -> bytes:
    # 資料日期在 fund_box 之外（與實站一致），驗證全頁 fallback 搜尋
    return """<html><body>
    <div class="fund_box p2">
      <ul>
        <li>基金淨資產(新台幣) 735,166,074</li>
        <li>基金在外流通單位數(單位) 73,729,000</li>
        <li>基金每單位淨值(新台幣) 9.9712</li>
      </ul>
    </div>
    <p>資料日期：2026/07/08</p>
    </body></html>""".encode("utf-8")


def _make_jpm_xlsx_with_summary() -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(
        [
            "Record Type",
            "Fund Ticker",
            "Estimated Total Market Value",
            "Estimated NAV",
            "Estimated NAV per Share",
            "Outstanding Shares",
        ]
    )
    ws.append(
        [
            "S",
            "00401A TW",
            3_170_140_850.0,
            3_230_570_130.84425,
            13.41069815,
            240_895_000,
        ]
    )
    ws.append([])
    ws.append(
        [
            "Record Type",
            "Constituent Ticker",
            "Constituent Description",
            "Shares or PAR Amount",
            "Market Value Base",
        ]
    )
    ws.append(["D", "2330", "TSMC", 10_000, 100_000.0])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── XLSX asset parsers (uni / fhtrust) ─────────────────────────────────────────


class TestUniFhtrustAssets(unittest.TestCase):
    def test_parse_uni_assets(self) -> None:
        from ETF.scrapers.official_api_scraper import _xlsx_raw_rows, _parse_uni_assets

        rows = _xlsx_raw_rows(_make_uni_xlsx())
        assets = _parse_uni_assets(rows)
        self.assertIsNotNone(assets)
        self.assertAlmostEqual(assets["aum"], 274_753_577_961)
        self.assertAlmostEqual(assets["units"], 9_199_209_000)
        self.assertAlmostEqual(assets["nav"], 29.87)
        self.assertEqual(assets["nav_date"], "2026-07-09")  # 民國115 → 西元2026

    def test_parse_fhtrust_assets(self) -> None:
        from ETF.scrapers.official_api_scraper import (
            _xlsx_raw_rows,
            _parse_fhtrust_assets,
        )

        rows = _xlsx_raw_rows(_make_fhtrust_xlsx())
        assets = _parse_fhtrust_assets(rows)
        self.assertIsNotNone(assets)
        self.assertAlmostEqual(assets["aum"], 75_346_169_965)
        self.assertAlmostEqual(assets["units"], 3_882_416_000)
        self.assertAlmostEqual(assets["nav"], 19.41)
        self.assertEqual(assets["nav_date"], "2026-07-09")

    def test_fetch_uni_returns_tuple_with_assets(self) -> None:
        from ETF.scrapers.official_api_scraper import _fetch_uni

        with patch(
            "ETF.scrapers.official_api_scraper._get", return_value=_make_uni_xlsx()
        ):
            holdings, assets = _fetch_uni("49YTW")
        self.assertGreater(len(holdings), 0)
        self.assertIsNotNone(assets)
        self.assertAlmostEqual(assets["nav"], 29.87)

    def test_fetch_fhtrust_returns_tuple_with_assets(self) -> None:
        from ETF.scrapers.official_api_scraper import _fetch_fhtrust

        with patch(
            "ETF.scrapers.official_api_scraper._get",
            return_value=_make_fhtrust_xlsx(),
        ):
            holdings, assets = _fetch_fhtrust("ETF23", "20260709")
        self.assertGreater(len(holdings), 0)
        self.assertIsNotNone(assets)
        self.assertAlmostEqual(assets["nav"], 19.41)


# ── Yuanta lightweight NAV fetch ────────────────────────────────────────────────


class TestYuantaAssets(unittest.TestCase):
    def test_fetch_yuanta_assets_parses_nav(self) -> None:
        from ETF.scrapers.official_api_scraper import _fetch_yuanta_assets

        html = """<html><body>
        <div class="td" data-v-1b37363e>基金淨資產價值 NTD $42,383,582,037.00</div>
        <div class="td" data-v-1b37363e>每受益權單位淨資產價值(元) NTD $17.73</div>
        <div class="td" data-v-1b37363e>已發行受益權單位總數 2,390,522,000</div>
        <div>公告日期：2026/07/10</div>
        </body></html>""".encode("utf-8")
        with patch("ETF.scrapers.official_api_scraper._get", return_value=html):
            assets = _fetch_yuanta_assets("00990A")
        self.assertIsNotNone(assets)
        self.assertAlmostEqual(assets["nav"], 17.73)
        self.assertAlmostEqual(assets["units"], 2_390_522_000)

    def test_fetch_yuanta_assets_returns_none_on_failure(self) -> None:
        from ETF.scrapers.official_api_scraper import _fetch_yuanta_assets

        with patch(
            "ETF.scrapers.official_api_scraper._get", side_effect=RuntimeError("boom")
        ):
            assets = _fetch_yuanta_assets("00990A")
        self.assertIsNone(assets)


# ── Taishin AJAX NAV endpoint ────────────────────────────────────────────────────


class TestTaishinAssets(unittest.TestCase):
    def test_fetch_taishin_assets_parses_latest_row(self) -> None:
        from ETF.scrapers.official_api_scraper import _fetch_taishin_assets

        html = """<table class="table table-striped listNo">
        <tr><td>2026-07-09</td><td>16.06</td><td>16.06</td><td>0%</td></tr>
        <tr><td>2026-07-08</td><td>15.99</td><td>15.94</td><td>-0.31%</td></tr>
        </table>""".encode("utf-8")
        mock_resp = MagicMock()
        mock_resp.read.return_value = html
        mock_resp.__enter__ = lambda s: s
        mock_resp.__exit__ = MagicMock(return_value=False)
        with patch("urllib.request.urlopen", return_value=mock_resp):
            assets = _fetch_taishin_assets("00987A", "20260709")
        self.assertIsNotNone(assets)
        self.assertAlmostEqual(assets["nav"], 16.06)
        self.assertEqual(assets["nav_date"], "2026-07-09")
        self.assertIsNone(assets["aum"])
        self.assertIsNone(assets["units"])


# ── Mega #asset_div (with HTML-comment decoy) ────────────────────────────────────


class TestMegaAssets(unittest.TestCase):
    def test_parse_mega_assets_strips_comment_decoy(self) -> None:
        from ETF.scrapers.official_api_scraper import _parse_mega_assets

        assets = _parse_mega_assets(_make_mega_html(with_comment_decoy=True))
        self.assertIsNotNone(assets)
        self.assertAlmostEqual(assets["aum"], 4_848_922_587)
        self.assertAlmostEqual(assets["units"], 321_899_000)
        self.assertAlmostEqual(assets["nav"], 15.06)
        self.assertEqual(assets["nav_date"], "2026-07-09")

    def test_fetch_mega_returns_tuple_with_assets(self) -> None:
        from ETF.scrapers.official_api_scraper import _fetch_mega

        html = _make_mega_html(with_comment_decoy=True).encode("utf-8")
        with patch("ETF.scrapers.official_api_scraper._get", return_value=html):
            holdings, assets = _fetch_mega("23")
        self.assertEqual(holdings, [])  # no fund-info rows in fixture
        self.assertIsNotNone(assets)
        self.assertAlmostEqual(assets["nav"], 15.06)


# ── ctbc_html Label_AUM02-04 ──────────────────────────────────────────────────────


class TestCtbcHtmlAssets(unittest.TestCase):
    def test_fetch_ctbc_html_returns_assets(self) -> None:
        from ETF.scrapers.official_api_scraper import _fetch_ctbc_html

        with patch(
            "ETF.scrapers.official_api_scraper._get",
            return_value=_make_ctbc_html_with_assets(),
        ):
            holdings, assets = _fetch_ctbc_html("00983A")
        self.assertGreater(len(holdings), 0)
        self.assertIsNotNone(assets)
        self.assertAlmostEqual(assets["aum"], 1_297_629_321)
        self.assertAlmostEqual(assets["units"], 105_686_000)
        self.assertAlmostEqual(assets["nav"], 12.28)
        self.assertEqual(assets["nav_date"], "2026/07/08")


# ── Fubon <li> assets (independent of holdings section) ──────────────────────────


class TestFubonAssets(unittest.TestCase):
    def test_fetch_fubon_returns_assets_even_without_holdings(self) -> None:
        from ETF.scrapers.official_api_scraper import _fetch_fubon

        with patch(
            "ETF.scrapers.official_api_scraper._get",
            return_value=_make_fubon_html(),
        ):
            holdings, assets = _fetch_fubon("00982D")
        self.assertEqual(holdings, [])  # bond ETF, no 持股明細 section in fixture
        self.assertIsNotNone(assets)
        self.assertAlmostEqual(assets["aum"], 735_166_074)
        self.assertAlmostEqual(assets["units"], 73_729_000)
        self.assertAlmostEqual(assets["nav"], 9.9712)
        self.assertEqual(assets["nav_date"], "2026-07-08")

    def test_fetch_fubon_uses_ssl_no_verify(self) -> None:
        """富邦憑證鏈缺 SKI，_get 必須帶 verify_ssl=False。"""
        from ETF.scrapers.official_api_scraper import _fetch_fubon

        captured: list[dict] = []

        def fake_get(url: str, **kwargs: object) -> bytes:
            captured.append({"url": url, "verify_ssl": kwargs.get("verify_ssl")})
            return _make_fubon_html()

        with patch("ETF.scrapers.official_api_scraper._get", side_effect=fake_get):
            _fetch_fubon("00982D")
        self.assertTrue(any(c["verify_ssl"] is False for c in captured))

    def test_fetch_holdings_attaches_assets_on_empty_holdings(self) -> None:
        """空持股（債券型）時 fetch_holdings 仍須把 fund_assets 附在空 df.attrs。"""
        from ETF.scrapers.official_api_scraper import fetch_holdings

        with patch(
            "ETF.scrapers.official_api_scraper._get",
            return_value=_make_fubon_html(),
        ):
            df = fetch_holdings("00982D")
        self.assertTrue(df.empty)
        assets = df.attrs.get("fund_assets")
        self.assertIsNotNone(assets)
        self.assertAlmostEqual(assets["nav"], 9.9712)


# ── JPM summary-row NAV/AUM/units ─────────────────────────────────────────────────


class TestJpmAssets(unittest.TestCase):
    def _fetch_jpm(self, xlsx_bytes: bytes):
        from ETF.scrapers.official_api_scraper import _fetch_jpm

        with patch("urllib.request.urlopen") as mock_urlopen:
            mock_resp = MagicMock()
            mock_resp.read.return_value = xlsx_bytes
            mock_resp.__enter__ = lambda s: s
            mock_resp.__exit__ = MagicMock(return_value=False)
            mock_urlopen.return_value = mock_resp
            return _fetch_jpm("https://am.jpmorgan.com/fake.xlsx")

    def test_fetch_jpm_returns_assets(self) -> None:
        holdings, assets = self._fetch_jpm(_make_jpm_xlsx_with_summary())
        self.assertGreater(len(holdings), 0)
        self.assertIsNotNone(assets)
        self.assertAlmostEqual(assets["nav"], 13.41069815)
        self.assertAlmostEqual(assets["units"], 240_895_000)
        # nav_date 恆為 None（Valuation Date 欄位人工複核發現不可靠，見函式 docstring）
        self.assertIsNone(assets["nav_date"])


# ── aum≈nav*units consistency guard ───────────────────────────────────────────────


class TestConsistencyCheck(unittest.TestCase):
    def test_consistent_values_pass_through(self) -> None:
        from ETF.scrapers.official_api_scraper import (
            _check_aum_nav_units_consistency,
            _fund_assets,
        )

        assets = _fund_assets(aum=1000.0, nav=10.0, units=100.0)
        result = _check_aum_nav_units_consistency(assets)
        self.assertEqual(result["aum"], 1000.0)
        self.assertEqual(result["units"], 100.0)

    def test_inconsistent_values_drop_aum_and_units(self) -> None:
        from ETF.scrapers.official_api_scraper import (
            _check_aum_nav_units_consistency,
            _fund_assets,
        )

        # aum way off from nav*units (10.0*100.0=1000.0, but aum=5000.0 → 400% err)
        assets = _fund_assets(aum=5000.0, nav=10.0, units=100.0)
        result = _check_aum_nav_units_consistency(assets)
        self.assertIsNone(result["aum"])
        self.assertIsNone(result["units"])
        self.assertEqual(result["nav"], 10.0)

    def test_missing_fields_pass_through_unchanged(self) -> None:
        from ETF.scrapers.official_api_scraper import (
            _check_aum_nav_units_consistency,
            _fund_assets,
        )

        assets = _fund_assets(nav=10.0)
        result = _check_aum_nav_units_consistency(assets)
        self.assertEqual(result, assets)

    def test_none_assets_passthrough(self) -> None:
        from ETF.scrapers.official_api_scraper import _check_aum_nav_units_consistency

        self.assertIsNone(_check_aum_nav_units_consistency(None))


if __name__ == "__main__":
    unittest.main()
